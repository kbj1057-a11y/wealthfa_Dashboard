"""
각 보험사 엑셀 파일의 시트 구조를 파악하는 스크립트 (ASCII only)
"""
import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

folder_path = r"G:\내 드라이브\안티그래비티\TEST\수수료분석\260227수수료"

for filename in sorted(os.listdir(folder_path)):
    if filename.endswith(".xlsx"):
        filepath = os.path.join(folder_path, filename)
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheets = wb.sheetnames
            print(f"\n[FILE] {filename}")
            for sh in sheets:
                ws = wb[sh]
                print(f"  [SHEET] {sh}  rows={ws.max_row}, cols={ws.max_column}")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 2:
                        break
                    print(f"    row{i+1}: {row[:8]}")
            wb.close()
        except Exception as e:
            print(f"[ERROR] {filename}: {e}")
