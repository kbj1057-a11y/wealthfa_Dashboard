"""
수수료분석 - 260227수수료 폴더의 모든 엑셀 파일을 올바르게 병합하는 스크립트

[문제 원인]
- 각 보험사 파일마다 25열 이후 수수료 항목 컬럼이 완전히 다름 (총 124개 고유 컬럼)
- 단순 이어붙이기 시 컬럼 불일치가 발생

[해결 전략]
- 공통 컬럼 (1~24): 번호, 적용년월, ... 보험료 → 그대로 사용
- 수수료 컬럼 (25~): 모든 파일의 수수료 컬럼을 수집하여 전체 컬럼 헤더 구성
- 각 행을 컬럼명 기반 딕셔너리로 변환 → 없는 컬럼은 None(빈 칸)으로 채움
- 마지막 '수수료계' 컬럼을 항상 맨 끝에 배치
"""
import os
import sys
import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────
BASE_DIR    = r"G:\내 드라이브\안티그래비티\TEST\수수료분석"
SOURCE_DIR  = os.path.join(BASE_DIR, "260227수수료")
OUTPUT_FILE = os.path.join(BASE_DIR, "260227수수료_통합_v2.xlsx")

# ─────────────────────────────────────────
# 1단계: 모든 파일의 컬럼 정보 수집
# ─────────────────────────────────────────
excel_files = sorted([
    f for f in os.listdir(SOURCE_DIR)
    if f.endswith(".xlsx") and not f.startswith("~$")
])

print(f"[1단계] 파일 수집: {len(excel_files)}개")

# 공통 컬럼 (1~24번째: 보험료까지) - 첫 번째 파일 기준
COMMON_COL_COUNT = 24  # 번호 ~ 보험료 (1~24번 컬럼)

# 각 파일의 헤더 구조 저장
file_info = {}  # filename -> {"headers": [...], "data_range": "보험료 이후 수수료 컬럼들"}

# 수수료 고유 컬럼 수집 (순서 보존을 위해 리스트 + set 병용)
fee_columns_ordered = []   # 등장 순서대로
fee_columns_seen   = set()
common_headers     = None

for filename in excel_files:
    filepath = os.path.join(SOURCE_DIR, filename)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # 헤더 행 읽기
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h) if h is not None else None for h in row]
        break

    # 공통 컬럼 검증 (첫 파일 기준으로 설정)
    if common_headers is None:
        common_headers = headers[:COMMON_COL_COUNT]
        print(f"\n[공통 컬럼 ({COMMON_COL_COUNT}개)]: {common_headers}")

    # 25번 이후 수수료 컬럼 수집 ('수수료계' 제외하고 수집)
    fee_cols_in_file = []
    for h in headers[COMMON_COL_COUNT:]:
        if h is None:
            continue
        if h == "수수료계":
            continue  # 마지막에 따로 배치
        fee_cols_in_file.append(h)
        if h not in fee_columns_seen:
            fee_columns_ordered.append(h)
            fee_columns_seen.add(h)

    file_info[filename] = {"headers": headers, "fee_cols": fee_cols_in_file}
    print(f"  {filename}: 수수료 컬럼 {len(fee_cols_in_file)}개")
    wb.close()

# 최종 통합 헤더: 공통 컬럼 + 전체 수수료 컬럼(순서 보존) + 수수료계
final_headers = common_headers + fee_columns_ordered + ["수수료계"]
print(f"\n[최종 통합 헤더] 총 {len(final_headers)}개 컬럼")
print(f"  공통: {COMMON_COL_COUNT}개 | 수수료: {len(fee_columns_ordered)}개 | 수수료계: 1개")

# ─────────────────────────────────────────
# 2단계: 통합 워크북 생성 & 데이터 기록
# ─────────────────────────────────────────
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "통합수수료_202602"

# 헤더 기록
ws_out.append(final_headers)

total_rows = 0
company_summary = []

for filename in excel_files:
    filepath = os.path.join(SOURCE_DIR, filename)
    wb_in = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws_in = wb_in.active

    # 이 파일의 헤더 → 인덱스 매핑
    file_headers = file_info[filename]["headers"]
    col_index_map = {h: i for i, h in enumerate(file_headers) if h is not None}

    rows_written = 0
    for row_idx, row_data in enumerate(ws_in.iter_rows(values_only=True)):
        if row_idx == 0:
            continue  # 헤더 스킵
        if all(v is None for v in row_data):
            continue  # 빈 행 스킵

        # 딕셔너리로 변환 (컬럼명 → 값)
        row_dict = {}
        for col_name, col_i in col_index_map.items():
            if col_i < len(row_data):
                row_dict[col_name] = row_data[col_i]

        # 최종 헤더 순서에 맞춰 행 생성 (없는 컬럼은 None)
        new_row = [row_dict.get(h, None) for h in final_headers]
        ws_out.append(new_row)
        rows_written += 1

    company_name = filename.replace("(사업부)수입수수료산출_", "").replace("_202602.xlsx", "")
    company_summary.append((company_name, rows_written))
    total_rows += rows_written
    print(f"  [{company_name}] {rows_written}행 기록 완료")
    wb_in.close()

# ─────────────────────────────────────────
# 3단계: 스타일 적용
# ─────────────────────────────────────────
print("\n[3단계] 스타일 적용 중...")

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 헤더 스타일
COMMON_FILL  = PatternFill("solid", fgColor="1F3864")  # 공통 컬럼: 진한 파랑
FEE_FILL     = PatternFill("solid", fgColor="1B6B8A")   # 수수료 컬럼: 다른 파랑
TOTAL_FILL   = PatternFill("solid", fgColor="8B0000")   # 수수료계: 진한 빨강
header_font  = Font(bold=True, color="FFFFFF", size=9)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, cell in enumerate(ws_out[1], start=1):
    if col_idx <= COMMON_COL_COUNT:
        cell.fill = COMMON_FILL
    elif col_idx == len(final_headers):
        cell.fill = TOTAL_FILL
    else:
        cell.fill = FEE_FILL
    cell.font      = header_font
    cell.alignment = header_align
    cell.border    = border

ws_out.row_dimensions[1].height = 32

# 데이터 행 스타일 (교차 줄무늬)
even_fill  = PatternFill("solid", fgColor="F5F8FC")
data_font  = Font(size=8)
data_align = Alignment(vertical="center")

for row_idx, row in enumerate(ws_out.iter_rows(min_row=2), start=2):
    bg = even_fill if row_idx % 2 == 0 else None
    for cell in row:
        cell.font      = data_font
        cell.alignment = data_align
        cell.border    = border
        if bg:
            cell.fill = bg

# 열 너비 자동 조정 (최대 20, 최소 6)
print("[3단계] 열 너비 조정 중 (컬럼 수가 많아 샘플 기반)...")
for col_idx in range(1, ws_out.max_column + 1):
    col_letter = get_column_letter(col_idx)
    header_val = str(ws_out[f"{col_letter}1"].value or "")
    # 헤더 길이 기준 (데이터 전체 순회는 컬럼 수가 많아 성능 부담)
    width = max(len(header_val) * 1.5 + 2, 6)
    width = min(width, 22)
    ws_out.column_dimensions[col_letter].width = width

# 틀 고정 (헤더 + 공통 컬럼 고정)
ws_out.freeze_panes = "Y2"  # 공통 25번째 열 이후 + 헤더 2행 고정

# ─────────────────────────────────────────
# 4단계: 저장
# ─────────────────────────────────────────
print(f"\n[4단계] 저장 중: {OUTPUT_FILE}")
wb_out.save(OUTPUT_FILE)

print(f"\n{'='*60}")
print(f"[완료] 통합 파일 저장: {OUTPUT_FILE}")
print(f"{'='*60}")
print(f"  총 데이터 행수 : {total_rows:,}행")
print(f"  총 컬럼 수     : {len(final_headers)}개")
print(f"    공통 컬럼    : {COMMON_COL_COUNT}개")
print(f"    수수료 컬럼  : {len(fee_columns_ordered)}개")
print(f"    수수료계     : 1개")
print(f"\n  [회사별 요약]")
for company, cnt in company_summary:
    print(f"  - {company:<25s}: {cnt:,}행")
print(f"{'='*60}")
