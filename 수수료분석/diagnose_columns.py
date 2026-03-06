"""
각 파일의 컬럼 구조를 상세 비교 진단하는 스크립트
- 컬럼명, 컬럼 수, 순서 불일치 여부를 모두 출력
"""
import os
import sys
import warnings
warnings.filterwarnings("ignore")

import openpyxl

sys.stdout = open(
    r"G:\내 드라이브\안티그래비티\TEST\수수료분석\diagnose_result.txt",
    "w", encoding="utf-8"
)

SOURCE_DIR = r"G:\내 드라이브\안티그래비티\TEST\수수료분석\260227수수료"

excel_files = sorted([
    f for f in os.listdir(SOURCE_DIR)
    if f.endswith(".xlsx") and not f.startswith("~$")
])

# 각 파일의 컬럼 분석
file_columns = {}  # filename -> [col1, col2, ...]

print("=" * 80)
print("【1단계】 각 파일의 컬럼 분석")
print("=" * 80)

for filename in excel_files:
    filepath = os.path.join(SOURCE_DIR, filename)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # 헤더 행(row 1) 읽기
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = list(row)
        break

    # 실제 데이터 최대 열 확인 (첫 10행 기준)
    max_actual_col = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10:
            break
        non_none = [j for j, v in enumerate(row) if v is not None]
        if non_none:
            max_actual_col = max(max_actual_col, max(non_none) + 1)

    # None이 아닌 헤더만 추출
    valid_headers = [(i, h) for i, h in enumerate(header_row) if h is not None]
    all_headers = header_row  # None 포함 전체

    file_columns[filename] = {
        "total_declared": len(header_row),
        "valid_count": len(valid_headers),
        "max_actual_col": max_actual_col,
        "headers": all_headers,
        "valid_headers": valid_headers
    }

    print(f"\n[파일] {filename}")
    print(f"  헤더 선언 컬럼 수 : {len(header_row)}")
    print(f"  유효 헤더(not None): {len(valid_headers)}")
    print(f"  실제 데이터 최대열: {max_actual_col}")
    print(f"  컬럼 목록:")
    for idx, (col_idx, col_name) in enumerate(valid_headers):
        print(f"    [{col_idx+1:02d}] {col_name}")

    # None 위치 확인
    none_positions = [i for i, h in enumerate(header_row) if h is None]
    if none_positions:
        print(f"  ** None 헤더 위치(0-based): {none_positions}")

    wb.close()

# ─────────────────────────────────────────
# 상호 비교
# ─────────────────────────────────────────
print("\n" + "=" * 80)
print("【2단계】 기준 파일 대비 컬럼 불일치 분석")
print("=" * 80)

# 첫 번째 파일을 기준으로
base_filename = excel_files[0]
base_headers = file_columns[base_filename]["headers"]
base_valid = [h for h in base_headers if h is not None]

print(f"\n기준 파일: {base_filename}")
print(f"기준 컬럼 ({len(base_valid)}개): {base_valid}")

for filename in excel_files[1:]:
    info = file_columns[filename]
    cur_headers = info["headers"]
    cur_valid = [h for h in cur_headers if h is not None]

    # 컬럼 수 차이
    count_match = len(cur_valid) == len(base_valid)
    # 순서 포함 완전 일치
    order_match = cur_valid == base_valid
    # 집합 기준 일치 (순서 무시)
    set_match = set(cur_valid) == set(base_valid)

    print(f"\n[비교] {filename}")
    print(f"  컬럼 수: {len(cur_valid)} (기준: {len(base_valid)}) -> {'OK' if count_match else 'MISMATCH!'}")
    print(f"  순서+내용 완전일치: {'YES' if order_match else 'NO - 차이 있음!'}")
    print(f"  집합 일치(순서무시): {'YES' if set_match else 'NO'}")

    if not order_match:
        # 차이 상세 출력
        missing_in_cur = [h for h in base_valid if h not in cur_valid]
        extra_in_cur   = [h for h in cur_valid if h not in base_valid]
        if missing_in_cur:
            print(f"  기준에는 있는데 이 파일엔 없는 컬럼: {missing_in_cur}")
        if extra_in_cur:
            print(f"  이 파일에만 있는 추가 컬럼: {extra_in_cur}")

        # 순서가 다른 컬럼 찾기
        order_diffs = []
        for i, (b, c) in enumerate(zip(base_valid, cur_valid[:len(base_valid)])):
            if b != c:
                order_diffs.append(f"  위치{i+1}: 기준='{b}' vs 현재='{c}'")
        if order_diffs:
            print(f"  순서 불일치:")
            for d in order_diffs:
                print(d)

print("\n\n" + "=" * 80)
print("【3단계】 전체 고유 컬럼 집합")
print("=" * 80)
all_cols = set()
for info in file_columns.values():
    for h in info["headers"]:
        if h is not None:
            all_cols.add(h)
print(f"전체 고유 컬럼 수: {len(all_cols)}")
for i, col in enumerate(sorted(all_cols), 1):
    print(f"  {i:02d}. {col}")

sys.stdout.close()
print("진단 완료 - diagnose_result.txt 저장됨", file=sys.stderr)
