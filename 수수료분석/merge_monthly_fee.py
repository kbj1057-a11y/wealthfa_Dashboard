"""
═══════════════════════════════════════════════════════════════
  수수료 월별 병합 스크립트  (범용 - 매월 재사용 가능)
  파일: merge_monthly_fee.py
═══════════════════════════════════════════════════════════════

[사용법]
  python merge_monthly_fee.py --yyyymm 202602

[처리 규칙]
  - 입력: 수수료분석/{YYYYMM}수수료/ 폴더의 모든 .xlsx
  - 출력: 수수료분석/{YYYYMM}수수료_생손보통합.xlsx  (2개 시트)
  - 시트1: 생명보험사 (파일명에 '생명' 포함)
  - 시트2: 손해보험사 (나머지)

[컬럼 구조]
  A~X  : 공통 기본 24개 (번호 ~ 보험료)
  Y    : 환산성적(생명) / 수정보험료(손해) ← 명칭 통일, 데이터 그대로
  Z    : 수수료계 ← 고정 위치
  AA~  : 각사 수수료 세부항목 (없는 회사 = 빈칸)

[Y열 명칭 통일]
  생명: 환산성적 / TP           → 환산성적
  손해: 수정보험료 / 신월정산수정P / 환산실적 → 수정보험료

═══════════════════════════════════════════════════════════════
"""

import os
import sys
import argparse
import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════
# ★ 설정 (새 보험사 추가 시 여기만 수정)
# ═══════════════════════════════════════

BASE_DIR         = r"G:\내 드라이브\안티그래비티\TEST\수수료분석"
COMMON_COL_COUNT = 24   # A(1)~X(24): 번호 ~ 보험료

# 생명보험사 Y열 후보 이름들 → 통일 이름
LIFE_Y_ALIASES  = {"환산성적", "TP"}
LIFE_UNIFIED_Y  = "환산성적"

# 손해보험사 Y열 후보 이름들 → 통일 이름
DAMAGE_Y_ALIASES = {"수정보험료", "신월정산수정P", "환산실적"}
DAMAGE_UNIFIED_Y = "수정보험료"

# 회사 분류: 파일명에 아래 키워드 포함 시 생명보험사
LIFE_KEYWORDS = ["생명"]

# ═══════════════════════════════════════
# 내부 함수
# ═══════════════════════════════════════

def col_letter(n):
    """숫자 → 엑셀 열 이름 (1=A, 26=Z, 27=AA, ...)"""
    letters = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        letters = chr(65 + r) + letters
    return letters


def is_life(filename):
    """파일명에 생명보험사 키워드 포함 여부"""
    return any(kw in filename for kw in LIFE_KEYWORDS)


def load_headers(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h) if h is not None else None for h in row]
        break
    wb.close()
    return headers


def build_group_info(file_list, y_aliases, unified_y):
    """
    파일 목록을 분석해 최종 헤더와 각사 컬럼 맵을 반환
    """
    all_headers    = {}
    fee_ordered    = []   # 세부항목 (등장 순서 보존)
    fee_seen       = set()

    for filepath, company in file_list:
        hdrs = load_headers(filepath)
        all_headers[company] = hdrs

        # 인덱스 25(=0-based 25)부터: Y열(index=24)과 수수료계 제외한 세부항목 수집
        for h in hdrs[COMMON_COL_COUNT + 1:]:
            if h is None or h == "수수료계":
                continue
            if h not in fee_seen:
                fee_ordered.append(h)
                fee_seen.add(h)

    # 최종 헤더: 공통24 + [Y통일명] + [수수료계] + [세부항목...]
    common = all_headers[file_list[0][1]][:COMMON_COL_COUNT]
    final_headers = common + [unified_y, "수수료계"] + fee_ordered

    # 각사 컬럼명→원본인덱스 맵
    col_maps = {
        company: {col: i for i, col in enumerate(hdrs) if col is not None}
        for (_, company), hdrs in [(pair, all_headers[pair[1]]) for pair in file_list]
    }

    return final_headers, col_maps, fee_ordered


def transform_row(row_data, col_map, y_aliases, final_headers):
    """원본 행 → 최종 헤더 순서에 맞춘 새 행"""
    row_dict = {col: row_data[idx] for col, idx in col_map.items() if idx < len(row_data)}

    new_row = []
    for col_name in final_headers:
        # Y열: 별칭 이름 중 하나로 저장된 값을 가져옴
        if col_name in (LIFE_UNIFIED_Y, DAMAGE_UNIFIED_Y):
            val = None
            for alias in y_aliases:
                if alias in row_dict:
                    val = row_dict[alias]
                    break
            new_row.append(val)
        else:
            new_row.append(row_dict.get(col_name))
    return new_row


def fill_sheet(ws, file_list, final_headers, col_maps, y_aliases):
    """시트에 헤더 + 데이터 기록"""
    ws.append(final_headers)
    total, summary = 0, []

    for filepath, company in file_list:
        wb_in = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws_in = wb_in.active
        count = 0

        for idx, row_data in enumerate(ws_in.iter_rows(values_only=True)):
            if idx == 0:
                continue
            if all(v is None for v in row_data):
                continue
            ws.append(transform_row(row_data, col_maps[company], y_aliases, final_headers))
            count += 1

        summary.append((company, count))
        total += count
        print(f"    [{company}] {count:,}행")
        wb_in.close()

    return total, summary


def apply_style(ws, common_count):
    """헤더 색상 + 데이터 교차줄무늬 + 열 너비 + 틀 고정"""
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLORS = {
        "common": "1F3864",   # A~X: 진한 파랑
        "y_col":  "2E75B6",   # Y: 파랑
        "total":  "C00000",   # Z(수수료계): 진한 빨강
        "detail": "375623",   # AA~: 초록
    }
    hdr_font  = Font(bold=True, color="FFFFFF", size=9)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, cell in enumerate(ws[1], start=1):
        if ci <= common_count:
            clr = COLORS["common"]
        elif ci == common_count + 1:
            clr = COLORS["y_col"]
        elif ci == common_count + 2:
            clr = COLORS["total"]
        else:
            clr = COLORS["detail"]
        cell.fill      = PatternFill("solid", fgColor=clr)
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = border

    ws.row_dimensions[1].height = 32

    # 데이터 교차 줄무늬
    even_fill  = PatternFill("solid", fgColor="F0F4F8")
    data_font  = Font(size=8)
    data_align = Alignment(vertical="center")
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = even_fill if ri % 2 == 0 else None
        for cell in row:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if bg:
                cell.fill = bg

    # 열 너비
    for ci in range(1, ws.max_column + 1):
        ltr = get_column_letter(ci)
        hval = str(ws[f"{ltr}1"].value or "")
        ws.column_dimensions[ltr].width = min(max(len(hval) * 1.4 + 2, 6), 22)

    # 틀 고정 (Y열 + 헤더 행 고정 → Z2부터 스크롤)
    ws.freeze_panes = "Z2"


# ═══════════════════════════════════════
# ■ 메인
# ═══════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="월별 수수료 생손보 통합 병합")
    parser.add_argument("--yyyymm", required=True, help="예: 202602")
    args = parser.parse_args()
    yyyymm = args.yyyymm

    source_dir  = os.path.join(BASE_DIR, f"{yyyymm}수수료")
    output_file = os.path.join(BASE_DIR, f"{yyyymm}수수료_생손보통합.xlsx")

    if not os.path.isdir(source_dir):
        print(f"[오류] 폴더 없음: {source_dir}")
        sys.exit(1)

    # 파일 분류
    all_xlsx = sorted([
        f for f in os.listdir(source_dir)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ])

    life_files   = [(os.path.join(source_dir, f), f.replace(f"(사업부)수입수수료산출_","").replace(f"_{yyyymm}.xlsx",""))
                    for f in all_xlsx if is_life(f)]
    damage_files = [(os.path.join(source_dir, f), f.replace(f"(사업부)수입수수료산출_","").replace(f"_{yyyymm}.xlsx",""))
                    for f in all_xlsx if not is_life(f)]

    print("=" * 60)
    print(f"  수수료 병합 시작: {yyyymm}")
    print(f"  입력 폴더: {source_dir}")
    print(f"  생명보험사: {len(life_files)}개  |  손해보험사: {len(damage_files)}개")
    print("=" * 60)

    wb_out = Workbook()

    # ── 생명보험사 시트 ──
    print(f"\n[1/2] 생명보험사 처리...")
    life_headers, life_maps, life_detail = build_group_info(life_files, LIFE_Y_ALIASES, LIFE_UNIFIED_Y)
    ws_life = wb_out.active
    ws_life.title = f"생명보험사_{yyyymm}"
    print(f"  헤더: 공통{COMMON_COL_COUNT} + Y(1) + 수수료계(1) + 세부({len(life_detail)}) = {len(life_headers)}개")
    life_total, life_sum = fill_sheet(ws_life, life_files, life_headers, life_maps, LIFE_Y_ALIASES)
    apply_style(ws_life, COMMON_COL_COUNT)

    # ── 손해보험사 시트 ──
    print(f"\n[2/2] 손해보험사 처리...")
    dmg_headers, dmg_maps, dmg_detail = build_group_info(damage_files, DAMAGE_Y_ALIASES, DAMAGE_UNIFIED_Y)
    ws_dmg = wb_out.create_sheet(f"손해보험사_{yyyymm}")
    print(f"  헤더: 공통{COMMON_COL_COUNT} + Y(1) + 수수료계(1) + 세부({len(dmg_detail)}) = {len(dmg_headers)}개")
    dmg_total, dmg_sum = fill_sheet(ws_dmg, damage_files, dmg_headers, dmg_maps, DAMAGE_Y_ALIASES)
    apply_style(ws_dmg, COMMON_COL_COUNT)

    # ── 저장 ──
    wb_out.save(output_file)

    # ── 결과 요약 ──
    print(f"\n{'=' * 60}")
    print(f"  완료! → {output_file}")
    print(f"{'=' * 60}")
    print(f"\n  [생명보험사] {len(life_headers)}컬럼 / {life_total:,}행")
    for c, n in life_sum:
        print(f"      {c:<20}: {n:,}행")
    print(f"\n  [손해보험사] {len(dmg_headers)}컬럼 / {dmg_total:,}행")
    for c, n in dmg_sum:
        print(f"      {c:<20}: {n:,}행")
    print(f"\n  전체 합계: {life_total + dmg_total:,}행")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
