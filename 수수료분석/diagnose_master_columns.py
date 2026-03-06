"""
계약 파일의 실제 컬럼 구조 확인
- 14번 인덱스(CONTRACT_ID_COL_IDX=14)가 정말 증권번호인지
- 마스터 파일의 증권번호 컬럼 위치도 확인
"""
import os, sys, warnings
warnings.filterwarnings("ignore")
import openpyxl, glob

CONTRACT_DIR = r"G:\내 드라이브\안티그래비티\TEST\계약관리(일자별)"
MASTER_FILE  = r"G:\내 드라이브\안티그래비티\TEST\매일업데이트\26년업적,수수료통계.xlsx"
FEE_DIR      = r"G:\내 드라이브\안티그래비티\TEST\수수료관리(일자별)"

out = []

# ─── 계약 파일 (가장 최신) ───
contract_files = sorted(glob.glob(os.path.join(CONTRACT_DIR, "*.xlsx")))
contract_files = [f for f in contract_files if not os.path.basename(f).startswith("~$")]
latest_contract = max(contract_files, key=os.path.getmtime) if contract_files else None

out.append("=" * 70)
out.append("【1. 계약 파일 컬럼 전체 목록】")
out.append(f"  파일: {os.path.basename(latest_contract)}")
out.append("=" * 70)

if latest_contract:
    wb = openpyxl.load_workbook(latest_contract, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        break
    # 실제 데이터 1행도 확인
    data_row = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=2, values_only=True)):
        data_row = list(row)
        break
    wb.close()

    out.append(f"  전체 컬럼 수: {len(headers)}\n")
    for i, (h, d) in enumerate(zip(headers, data_row)):
        marker = ""
        if str(h) == "증권번호" or (d and len(str(d)) > 8 and str(d).replace('-','').isdigit()):
            marker = "  ← 증권번호 후보"
        if i == 14:
            marker += "  ← 현재 스크립트가 사용 중인 INDEX=14"
        out.append(f"  [{i:02d}] {str(h):<30} | 샘플: {str(d)[:25]}{marker}")

# ─── 마스터 파일 컬럼 ───
out.append("\n" + "=" * 70)
out.append("【2. 마스터 파일(RAWDATA 시트) 컬럼】")
out.append("=" * 70)

wb_m = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
ws_m = wb_m["RAWDATA"]
master_headers = []
master_data = []
for i_r, row in enumerate(ws_m.iter_rows(min_row=1, max_row=2, values_only=True)):
    if i_r == 0:
        master_headers = list(row)
    else:
        master_data = list(row)
wb_m.close()

out.append(f"  전체 컬럼 수: {len(master_headers)}\n")
for i, (h, d) in enumerate(zip(master_headers, master_data)):
    marker = "  ← 증권번호(KEY)" if str(h) == "증권번호" else ""
    out.append(f"  [{i:02d}] {str(h):<30} | 샘플: {str(d)[:25]}{marker}")

# ─── 수수료 파일 컬럼 ───
fee_files = sorted(glob.glob(os.path.join(FEE_DIR, "*.xlsx")))
fee_files = [f for f in fee_files if not os.path.basename(f).startswith("~$")]
latest_fee = max(fee_files, key=os.path.getmtime) if fee_files else None

out.append("\n" + "=" * 70)
out.append("【3. 수수료 파일 컬럼】")
out.append(f"  파일: {os.path.basename(latest_fee)}")
out.append("=" * 70)

if latest_fee:
    wb_f = openpyxl.load_workbook(latest_fee, read_only=True, data_only=True)
    ws_f = wb_f.active
    fee_headers = []
    fee_data = []
    for i_r, row in enumerate(ws_f.iter_rows(min_row=1, max_row=2, values_only=True)):
        if i_r == 0:
            fee_headers = list(row)
        else:
            fee_data = list(row)
    wb_f.close()
    out.append(f"  전체 컬럼 수: {len(fee_headers)}\n")
    for i, (h, d) in enumerate(zip(fee_headers, fee_data)):
        marker = "  ← 증권번호(KEY)" if str(h) == "증권번호" else ""
        out.append(f"  [{i:02d}] {str(h):<30} | 샘플: {str(d)[:25]}{marker}")

# 저장
output_path = r"G:\내 드라이브\안티그래비티\TEST\수수료분석\diagnose_master_columns.txt"
with open(output_path, "w", encoding="utf-8") as f:
    f.write("\n".join(out))
print(f"저장 완료: {output_path}", file=sys.stderr)
