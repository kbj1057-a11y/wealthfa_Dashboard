"""
수수료 통합 병합 스크립트 v3 - 생명/손해보험사 분리 2시트

[컬럼 구조]
  A~X  (1~24)  : 공통 기본 24개 (번호~보험료) - 동일
  Y    (25)    : 생명=환산성적(통일), 손해=수정보험료(통일) ← 명칭만 통일
  Z    (26)    : 수수료계 ← 바로 다음 위치
  AA~  (27~)   : 각사 수수료 세부항목 (없는 회사 = 빈칸)

[Y열 명칭 매핑]
  생명: 환산성적/TP → 환산성적
  손해: 수정보험료/신월정산수정P/환산실적 → 수정보험료
"""
import os
import sys
import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# 경로 & 설정
# ─────────────────────────────────────────
SOURCE_DIR  = r"G:\내 드라이브\안티그래비티\TEST\수수료분석\260227수수료"
OUTPUT_FILE = r"G:\내 드라이브\안티그래비티\TEST\수수료분석\260227수수료_생손보통합_v3.xlsx"

COMMON_COL_COUNT = 24   # A~X: 번호~보험료

# Y열 원본명 → 통일명 매핑
LIFE_Y_ALIASES   = {"환산성적", "TP"}       # 이 이름들이 오면 Y열로 인식
LIFE_UNIFIED_Y   = "환산성적"               # 생명 통일 Y열 이름

DAMAGE_Y_ALIASES = {"수정보험료", "신월정산수정P", "환산실적"}  # 손해 Y열 후보
DAMAGE_UNIFIED_Y = "수정보험료"             # 손해 통일 Y열 이름

# 회사 목록
LIFE_FILES = [
    ("(사업부)수입수수료산출_농협생명_202602.xlsx",    "농협생명"),
    ("(사업부)수입수수료산출_미래에셋생명_202602.xlsx", "미래에셋생명"),
    ("(사업부)수입수수료산출_삼성생명_202602.xlsx",    "삼성생명"),
]
DAMAGE_FILES = [
    ("(사업부)수입수수료산출_AIG손해보험_202602.xlsx",  "AIG손해보험"),
    ("(사업부)수입수수료산출_DB손해보험_202602.xlsx",   "DB손해보험"),
    ("(사업부)수입수수료산출_KB손해보험_202602.xlsx",   "KB손해보험"),
    ("(사업부)수입수수료산출_롯데손해보험_202602.xlsx",  "롯데손해보험"),
    ("(사업부)수입수수료산출_메리츠화재_202602.xlsx",   "메리츠화재"),
    ("(사업부)수입수수료산출_삼성화재_202602.xlsx",    "삼성화재"),
    ("(사업부)수입수수료산출_하나손해_202602.xlsx",    "하나손해"),
    ("(사업부)수입수수료산출_한화손해보험_202602.xlsx",  "한화손해보험"),
    ("(사업부)수입수수료산출_현대해상_202602.xlsx",    "현대해상"),
    ("(사업부)수입수수료산출_흥국화재_202602.xlsx",    "흥국화재"),
]

# ─────────────────────────────────────────
# 헬퍼: 파일 헤더 로딩
# ─────────────────────────────────────────
def load_headers(filename):
    fp = os.path.join(SOURCE_DIR, filename)
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h) if h is not None else None for h in row]
        break
    wb.close()
    return headers

# ─────────────────────────────────────────
# 헬퍼: 그룹 분석 → 최종 헤더 & 각사 컬럼 맵 생성
# ─────────────────────────────────────────
def build_group_info(file_list, y_aliases, unified_y_name):
    """
    반환값:
      final_headers : [공통24] + [통일Y] + [수수료계] + [세부항목...]
      company_maps  : { company: { col_name: index_in_original } }
      y_col_index   : 원본 파일에서 Y열의 인덱스 (항상 24)
    """
    group_headers   = {}   # company → original header list
    fee_detail_ordered = [] # 세부항목 (등장순, 중복 없이)
    fee_detail_seen    = set()

    for filename, company in file_list:
        headers = load_headers(filename)
        group_headers[company] = headers

        # Y열(index 24)과 수수료계는 제외하고 세부항목만 수집
        for h in headers[COMMON_COL_COUNT + 1:]:   # 25번(index=25)부터
            if h is None or h == "수수료계":
                continue
            if h not in fee_detail_seen:
                fee_detail_ordered.append(h)
                fee_detail_seen.add(h)

    # 최종 헤더 구성: 공통24 + Y통일명 + 수수료계 + 세부항목들
    common_headers = group_headers[file_list[0][1]][:COMMON_COL_COUNT]
    final_headers  = common_headers + [unified_y_name, "수수료계"] + fee_detail_ordered

    # 각 회사의 원본 헤더→인덱스 맵
    company_maps = {}
    for _, company in file_list:
        h = group_headers[company]
        company_maps[company] = {col: i for i, col in enumerate(h) if col is not None}

    return final_headers, company_maps, fee_detail_ordered

# ─────────────────────────────────────────
# 헬퍼: 데이터 행 변환
# ─────────────────────────────────────────
def transform_row(row_data, col_map, y_aliases, final_headers):
    """
    원본 row_data → final_headers 순서에 맞춘 new_row 생성
    """
    row_dict = {col: row_data[idx] for col, idx in col_map.items() if idx < len(row_data)}

    new_row = []
    for col_name in final_headers:
        if col_name in ("환산성적", "수정보험료"):
            # Y열 통일: y_aliases 중 어떤 이름으로 저장됐든 값을 가져옴
            val = None
            for alias in y_aliases:
                if alias in row_dict:
                    val = row_dict[alias]
                    break
            new_row.append(val)
        else:
            new_row.append(row_dict.get(col_name, None))
    return new_row

# ─────────────────────────────────────────
# 헬퍼: 시트에 데이터 기록
# ─────────────────────────────────────────
def fill_sheet(ws, file_list, final_headers, company_maps, y_aliases, unified_y_name):
    ws.append(final_headers)   # 헤더 1행 기록

    total = 0
    summary = []

    for filename, company in file_list:
        fp = os.path.join(SOURCE_DIR, filename)
        wb_in = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws_in = wb_in.active
        col_map = company_maps[company]
        rows_written = 0

        for row_idx, row_data in enumerate(ws_in.iter_rows(values_only=True)):
            if row_idx == 0:
                continue
            if all(v is None for v in row_data):
                continue
            new_row = transform_row(row_data, col_map, y_aliases, final_headers)
            ws.append(new_row)
            rows_written += 1

        summary.append((company, rows_written))
        total += rows_written
        print(f"    [{company}] {rows_written:,}행 기록")
        wb_in.close()

    return total, summary

# ─────────────────────────────────────────
# 헬퍼: 시트 스타일 적용
# ─────────────────────────────────────────
def apply_style(ws, common_count, unified_y_name, detail_start_idx):
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더 색상 정의
    COLOR_COMMON = "1F3864"  # 진한 파랑 (공통)
    COLOR_Y      = "2E75B6"  # 파랑 (Y열 통일)
    COLOR_TOTAL  = "C00000"  # 진한 빨강 (수수료계)
    COLOR_DETAIL = "375623"  # 진한 초록 (세부항목)

    hdr_font  = Font(bold=True, color="FFFFFF", size=9)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, cell in enumerate(ws[1], start=1):
        if col_idx <= common_count:           # A~X 공통
            fill_color = COLOR_COMMON
        elif col_idx == common_count + 1:     # Y열
            fill_color = COLOR_Y
        elif col_idx == common_count + 2:     # Z열 (수수료계)
            fill_color = COLOR_TOTAL
        else:                                  # AA~ 세부항목
            fill_color = COLOR_DETAIL
        cell.fill      = PatternFill("solid", fgColor=fill_color)
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = border

    ws.row_dimensions[1].height = 32

    # 데이터 행 (교차 줄무늬)
    even_fill = PatternFill("solid", fgColor="F0F4F8")
    data_font  = Font(size=8)
    data_align = Alignment(vertical="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = even_fill if row_idx % 2 == 0 else None
        for cell in row:
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = border
            if bg:
                cell.fill = bg

    # 열 너비 (헤더 길이 기반)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_val = str(ws[f"{col_letter}1"].value or "")
        width = max(len(header_val) * 1.4 + 2, 6)
        ws.column_dimensions[col_letter].width = min(width, 22)

    # 틀 고정: Y열 + 헤더(1행) → Z열 2행부터 스크롤
    ws.freeze_panes = "Z2"

# ─────────────────────────────────────────
# ■ 메인 실행
# ─────────────────────────────────────────
print("=" * 60)
print("[시작] 생/손보 통합 병합 v3")
print("=" * 60)

wb_out = Workbook()

# ── 시트 1: 생명보험사 ──
print("\n[1/2] 생명보험사 시트 처리 중...")
life_headers, life_maps, life_detail = build_group_info(
    LIFE_FILES, LIFE_Y_ALIASES, LIFE_UNIFIED_Y
)
ws_life = wb_out.active
ws_life.title = "생명보험사_202602"

print(f"  최종 헤더: {len(life_headers)}개")
print(f"    공통24 + Y(환산성적) + 수수료계 + 세부항목{len(life_detail)}개")

life_total, life_summary = fill_sheet(
    ws_life, LIFE_FILES, life_headers, life_maps, LIFE_Y_ALIASES, LIFE_UNIFIED_Y
)
apply_style(ws_life, COMMON_COL_COUNT, LIFE_UNIFIED_Y, COMMON_COL_COUNT + 2)

# ── 시트 2: 손해보험사 ──
print("\n[2/2] 손해보험사 시트 처리 중...")
damage_headers, damage_maps, damage_detail = build_group_info(
    DAMAGE_FILES, DAMAGE_Y_ALIASES, DAMAGE_UNIFIED_Y
)
ws_damage = wb_out.create_sheet("손해보험사_202602")

print(f"  최종 헤더: {len(damage_headers)}개")
print(f"    공통24 + Y(수정보험료) + 수수료계 + 세부항목{len(damage_detail)}개")

damage_total, damage_summary = fill_sheet(
    ws_damage, DAMAGE_FILES, damage_headers, damage_maps, DAMAGE_Y_ALIASES, DAMAGE_UNIFIED_Y
)
apply_style(ws_damage, COMMON_COL_COUNT, DAMAGE_UNIFIED_Y, COMMON_COL_COUNT + 2)

# ── 저장 ──
print(f"\n[저장 중] {OUTPUT_FILE}")
wb_out.save(OUTPUT_FILE)

# ── 최종 결과 ──
print(f"\n{'=' * 60}")
print(f"[완료] {OUTPUT_FILE}")
print(f"{'=' * 60}")
print(f"\n  ■ 생명보험사 시트")
print(f"    총 컬럼 : {len(life_headers)}개  (공통24 + Y1 + 수수료계1 + 세부{len(life_detail)})")
print(f"    총 데이터: {life_total:,}행")
for c, n in life_summary:
    print(f"      - {c:<20}: {n:,}행")

print(f"\n  ■ 손해보험사 시트")
print(f"    총 컬럼 : {len(damage_headers)}개  (공통24 + Y1 + 수수료계1 + 세부{len(damage_detail)})")
print(f"    총 데이터: {damage_total:,}행")
for c, n in damage_summary:
    print(f"      - {c:<20}: {n:,}행")

print(f"\n  ■ 전체 합계: {life_total + damage_total:,}행")
print(f"{'=' * 60}")
