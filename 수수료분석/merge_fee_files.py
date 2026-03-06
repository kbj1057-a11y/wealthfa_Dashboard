"""
수수료분석 - 260227수수료 폴더의 모든 엑셀 파일을 하나로 병합하는 스크립트

전략:
- 모든 파일은 동일한 컬럼 구조를 가짐
- 첫 번째 파일의 헤더(row1)만 사용하고 나머지는 데이터만 이어붙임
- 출력 파일: 수수료분석/260227수수료_통합.xlsx
"""
import os
import sys
import warnings

# openpyxl 경고 억제
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────
BASE_DIR    = r"G:\내 드라이브\안티그래비티\TEST\수수료분석"
SOURCE_DIR  = os.path.join(BASE_DIR, "260227수수료")
OUTPUT_FILE = os.path.join(BASE_DIR, "260227수수료_통합.xlsx")

# ─────────────────────────────────────────
# 파일 목록 수집 (정렬)
# ─────────────────────────────────────────
excel_files = sorted([
    f for f in os.listdir(SOURCE_DIR)
    if f.endswith(".xlsx") and not f.startswith("~$")
])

print(f"[INFO] 발견된 파일 수: {len(excel_files)}")
for f in excel_files:
    print(f"  - {f}")

# ─────────────────────────────────────────
# 통합 워크북 생성
# ─────────────────────────────────────────
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "통합수수료_202602"

header_done = False     # 헤더를 한 번만 쓰기 위한 플래그
total_data_rows = 0
company_summary = []    # (회사명, 데이터행수) 요약

# ─────────────────────────────────────────
# 각 파일 순회 → 데이터 복사
# ─────────────────────────────────────────
for filename in excel_files:
    filepath = os.path.join(SOURCE_DIR, filename)
    print(f"\n[PROCESSING] {filename}")

    try:
        wb_in = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws_in = wb_in.active   # 시트가 1개이므로 active 사용

        rows_written = 0
        for row_idx, row in enumerate(ws_in.iter_rows(values_only=True)):
            # 첫 번째 행(헤더) 처리
            if row_idx == 0:
                if not header_done:
                    ws_out.append(list(row))
                    header_done = True
                    print(f"  -> 헤더 기록 완료: {len(row)}개 컬럼")
                # else: 헤더 스킵
                continue

            # 빈 행은 건너뜀
            if all(v is None for v in row):
                continue

            ws_out.append(list(row))
            rows_written += 1

        # 회사명 요약 (파일명에서 추출)
        company_name = filename.replace("(사업부)수입수수료산출_", "").replace("_202602.xlsx", "")
        company_summary.append((company_name, rows_written))
        total_data_rows += rows_written
        print(f"  -> 데이터 {rows_written}행 복사 완료")
        wb_in.close()

    except Exception as e:
        print(f"  [ERROR] {filename}: {e}")

# ─────────────────────────────────────────
# 헤더 스타일 적용 (진한 파란색 배경 + 흰 글씨)
# ─────────────────────────────────────────
header_fill  = PatternFill("solid", fgColor="1F3864")
header_font  = Font(bold=True, color="FFFFFF", size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws_out[1]:
    cell.fill  = header_fill
    cell.font  = header_font
    cell.alignment = header_align
    cell.border = border

ws_out.row_dimensions[1].height = 30

# ─────────────────────────────────────────
# 데이터 행 교차 줄무늬 (흰색/연회색)
# ─────────────────────────────────────────
light_fill = PatternFill("solid", fgColor="F2F2F2")
data_font  = Font(size=9)
data_align = Alignment(vertical="center")

for row_idx, row in enumerate(ws_out.iter_rows(min_row=2), start=2):
    bg = light_fill if row_idx % 2 == 0 else None
    for cell in row:
        cell.font      = data_font
        cell.alignment = data_align
        cell.border    = border
        if bg:
            cell.fill = bg

# ─────────────────────────────────────────
# 열 너비 자동 조정 (최대 40)
# ─────────────────────────────────────────
for col_idx in range(1, ws_out.max_column + 1):
    max_len = 0
    col_letter = get_column_letter(col_idx)
    for row in ws_out.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
        val = row[0]
        if val is not None:
            max_len = max(max_len, len(str(val)))
    adjusted_width = min(max_len + 4, 40)
    ws_out.column_dimensions[col_letter].width = adjusted_width

# 틀 고정 (헤더 행 고정)
ws_out.freeze_panes = "A2"

# ─────────────────────────────────────────
# 저장
# ─────────────────────────────────────────
wb_out.save(OUTPUT_FILE)
print(f"\n{'='*60}")
print(f"[완료] 통합 파일 저장: {OUTPUT_FILE}")
print(f"{'='*60}")
print(f"  총 데이터 행수 : {total_data_rows:,}행")
print(f"  컬럼 수        : {ws_out.max_column}개")
print(f"\n  [회사별 요약]")
for company, cnt in company_summary:
    print(f"  - {company:20s}: {cnt:,}행")
print(f"{'='*60}")
